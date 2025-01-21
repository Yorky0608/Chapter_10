import openpyxl as op
import ast

path = "Candy_Request_Responses_1.xlsx"

# Load the workbook and the active sheet
wb = op.load_workbook(path)
sheet = wb.active
max_col = sheet.max_column
max_row = sheet.max_row

# Initialize fav_candies dictionary
fav_candies = {}

# Read the existing candy data into fav_candies
for a in range(2, max_row + 1):
    cell1 = sheet.cell(row=a, column=2)
    cell2 = sheet.cell(row=a, column=3)

    cell1 = cell1.value.title().strip()
    if isinstance(cell2.value, str) and '[' in cell2.value:
        try:
            # Try to safely evaluate the string as a list
            cell2_value = ast.literal_eval(cell2.value)
        except (ValueError, SyntaxError):
            cell2_value = [cell2.value]
    else:
        cell2_value = [cell2.value]

    fav_candies[cell1] = cell2_value


# Function to arrange candy names alphabetically
def arrange_by_candies(fav_candies):
    candy_dict = {}
    for name, candies in fav_candies.items():
        for candy in candies:
            can = candy.replace(' ', '')  # Standardize candy name (remove spaces)
            if can not in candy_dict:
                candy_dict[can] = [name]
            else:
                candy_dict[can].append(name)
    return candy_dict


# Function to ask user for candy
def ask_for_candy():
    candy = input('What candy would you like to add? ')
    return candy.lower()


# Function to add candies to a person's list
def add_candies(cell1, cell2, index):
    ws = wb.active
    new_candy = ask_for_candy()

    # Ensure cell2 is treated as a list, not a Cell object
    if not isinstance(cell2.value, list):
        cell2_value = [cell2.value]
    else:
        cell2_value = cell2.value  # Already a list

    # Update the person's candy list
    cell2_value.append(new_candy)  # Add the new candy to the list
    fav_candies[cell1.value].append(new_candy)  # Update the dictionary as well

    # Save the updated list back into the cell (as a string representation of the list)
    ws.cell(row=index, column=3, value=str(cell2_value))  # Convert list to string to store in Excel


# Function to add a new person
def add_new_person(name, count):
    count += 1
    new_candy = [[f"Time{count}", name, ask_for_candy()]]
    for row in new_candy:
        sheet.append(row)


# Function to ask user for name and candy updates
def ask_add_candies(fav_candies, count, max_row):
    run = True
    while run:
        name = input('What is your name? ')
        try:
            int(name)
        except ValueError:
            run = False
        else:
            print('Please provide your name.')

    if name in fav_candies:
        while True:
            y_n = input('Would you like to add a candy to your list? y/n ')
            if y_n == 'y':
                for index in range(2, max_row + 1):
                    cell1 = sheet.cell(row=index, column=2)
                    cell2 = sheet.cell(row=index, column=3)

                    if cell1.value.title().strip() == name:
                        add_candies(cell1, cell2, index)
                        break
            elif y_n == 'n':
                break
            else:
                print('Please provide y for yes and n for no.')
    else:
        print('Your name is not in our database.')
        print('Adding you to our database.')
        add_new_person(name, count)


# Run the add-candy flow
ask_add_candies(fav_candies, 0, max_row)

# Save the workbook after updates
wb.save(path)

# Reload the updated sheet and show the results
wb = op.load_workbook(path)
sheet = wb.active
max_col = sheet.max_column
max_row = sheet.max_row

# Re-load the candies into the dictionary
fav_candies = {}
for a in range(2, max_row + 1):
    cell1 = sheet.cell(row=a, column=2)
    cell2 = sheet.cell(row=a, column=3)

    cell1 = cell1.value.title().strip()
    if isinstance(cell2.value, str) and '[' in cell2.value:
        try:
            cell2_value = ast.literal_eval(cell2.value)
        except (ValueError, SyntaxError):
            cell2_value = [cell2.value]
    else:
        cell2_value = [cell2.value]

    fav_candies[cell1] = cell2_value

print(fav_candies)
print(arrange_by_candies(fav_candies))
